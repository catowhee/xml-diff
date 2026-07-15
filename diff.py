import gzip
import io
import os
import re
import tempfile
import zipfile
from datetime import datetime
from urllib.request import urlopen

import paramiko
import pandas as pd
from dotenv import load_dotenv
from lxml import etree

load_dotenv()

CATALOG_NS = "http://www.demandware.com/xml/impex/catalog/2006-10-31"
PRODUCT_TAG = f"{{{CATALOG_NS}}}product"
VARIATIONS_TAG = f"{{{CATALOG_NS}}}variations"
IMAGES_TAG = f"{{{CATALOG_NS}}}images"
CUSTOM_ATTRS_TAG = f"{{{CATALOG_NS}}}custom-attributes"

SFTP_FULL_PATHS = [
    "/incoming/fdx_fulls/style.xml.gz",
    "/incoming/fdx_fulls/style-colour.xml.gz",
    "/incoming/fdx_fulls/sku.xml.gz",
]

PRODUCT_LEVELS = ["style", "style-colour", "sku"]

SKIP_TAGS = {VARIATIONS_TAG, IMAGES_TAG, CUSTOM_ATTRS_TAG}


# ---------------------------------------------------------------------------
# SFTP / file loading
# ---------------------------------------------------------------------------

def _sftp_client():
    host = os.environ["SFTP_HOST"]
    port = int(os.environ.get("SFTP_PORT", 22))
    username = os.environ["SFTP_USERNAME"]
    password = os.environ["SFTP_PASSWORD"]
    transport = paramiko.Transport((host, port))
    transport.connect(username=username, password=password)
    return paramiko.SFTPClient.from_transport(transport), transport


def download_from_sftp(remote_path):
    """Download a gzipped XML from SFTP and return a BytesIO of the decompressed XML."""
    print(f"  Downloading {remote_path} ...")
    transport = paramiko.Transport((os.environ["SFTP_HOST"], int(os.environ.get("SFTP_PORT", 22))))
    transport.connect(username=os.environ["SFTP_USERNAME"], password=os.environ["SFTP_PASSWORD"])
    sftp = paramiko.SFTPClient.from_transport(transport)
    try:
        buf = io.BytesIO()
        sftp.getfo(remote_path, buf)
        buf.seek(0)
    finally:
        sftp.close()
        transport.close()

    with gzip.open(buf) as gz:
        return io.BytesIO(gz.read())


def iter_xml_buffers_from_zip(z):
    """Recursively yield (name, BytesIO) for every XML inside a zip (one level deep)."""
    for name in z.namelist():
        if name.lower().endswith(".xml"):
            yield name, io.BytesIO(z.read(name))
        elif name.lower().endswith(".zip"):
            with z.open(name) as inner_file:
                with zipfile.ZipFile(io.BytesIO(inner_file.read())) as inner_zip:
                    yield from iter_xml_buffers_from_zip(inner_zip)


def get_xml_buffers(source):
    """Return a list of (name, BytesIO) from a zip path or a raw XML BytesIO."""
    if isinstance(source, str):
        with zipfile.ZipFile(source) as z:
            return list(iter_xml_buffers_from_zip(z))
    else:
        name = os.path.basename(os.environ.get("SFTP_REMOTE_PATH", "sftp.xml"))
        if name.endswith(".gz"):
            name = name[:-3]
        return [(name, source)]


# ---------------------------------------------------------------------------
# XML parsing
# ---------------------------------------------------------------------------

def fetch_schema(buf):
    for _, el in etree.iterparse(buf, events=("start",)):
        schema_location = el.get("{http://www.w3.org/2001/XMLSchema-instance}schemaLocation")
        if not schema_location:
            return None
        parts = schema_location.split()
        urls = parts[1::2]
        schema_doc = etree.parse(io.BytesIO(urlopen(urls[0]).read()))
        return etree.XMLSchema(schema_doc)
    return None


_HTML_COMMENT = re.compile(r"<!--.*?-->", re.DOTALL)


def clean_long_description(text):
    if not text:
        return text
    text = _HTML_COMMENT.sub("", text)
    text = re.sub(r">\s+<", "><", text)   # whitespace between tags
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


_SKU_RE = re.compile(r"^\d{13,}-")


def product_level(el, product_id):
    if el.find(VARIATIONS_TAG) is not None:
        return "style"
    if _SKU_RE.match(product_id):
        return "sku"
    if "-" in product_id:
        return "style-colour"
    return "sku"


def _strip_namespaces(el):
    new_el = etree.Element(etree.QName(el).localname)
    new_el.text = el.text
    new_el.tail = el.tail
    for child in el:
        new_el.append(_strip_namespaces(child))
    return new_el


def custom_attr_value(el):
    children = list(el)
    if children:
        return "".join(etree.tostring(_strip_namespaces(c), encoding="unicode") for c in children)
    return el.text if el.text and el.text.strip() else None


def extract_product(el, product_id):
    row = {"product-id": product_id, "product-level": product_level(el, product_id)}

    for child in el:
        if child.tag in SKIP_TAGS:
            continue
        field = etree.QName(child.tag).localname
        site_id = child.get("site-id")
        col = f"{field}.{site_id}" if site_id else field
        if col not in row:
            value = child.text
            if field == "long-description":
                value = clean_long_description(value)
            row[col] = value

    custom_attrs = el.find(CUSTOM_ATTRS_TAG)
    if custom_attrs is not None:
        for attr in custom_attrs:
            attribute_id = attr.get("attribute-id")
            site_id = attr.get("site-id")
            col = f"custom.{attribute_id}.{site_id}" if site_id else f"custom.{attribute_id}"
            row[col] = custom_attr_value(attr)

    variations = el.find(VARIATIONS_TAG)
    if variations is not None:
        attrs_el = variations.find(f"{{{CATALOG_NS}}}attributes")
        if attrs_el is not None:
            shared = attrs_el.findall(f"{{{CATALOG_NS}}}shared-variation-attribute")
            row["variation-attribute-ids"] = "|".join(a.get("variation-attribute-id") for a in shared)
            row["variation-attribute-names"] = "|".join(a.get("attribute-id") for a in shared)

        variants_el = variations.find(f"{{{CATALOG_NS}}}variants")
        if variants_el is not None:
            row["variants"] = "|".join(sorted(
                v.get("product-id") for v in variants_el.findall(f"{{{CATALOG_NS}}}variant")
            ))

        groups_el = variations.find(f"{{{CATALOG_NS}}}variation-groups")
        if groups_el is not None:
            row["variation-groups"] = "|".join(sorted(
                g.get("product-id") for g in groups_el.findall(f"{{{CATALOG_NS}}}variation-group")
            ))

    return row


def stream_products(buf, schema=None):
    for _, el in etree.iterparse(buf, events=("end",), tag=PRODUCT_TAG, schema=schema):
        product_id = el.get("product-id")
        if product_id:
            yield extract_product(el, product_id)
        el.clear()


def load_table(xmls, label=""):
    """Parse a list of (name, BytesIO) into a single DataFrame."""
    chunks = []
    for name, buf in xmls:
        print(f"  {name}")
        schema = fetch_schema(buf)
        if schema is None:
            print("    [warn] No xsi:schemaLocation, skipping validation")
        buf.seek(0)
        try:
            chunks.append(pd.DataFrame(stream_products(buf, schema=schema)))
        except etree.XMLSyntaxError as e:
            print(f"    [error] {e}")
    if not chunks:
        return pd.DataFrame()
    df = pd.concat(chunks, ignore_index=True)
    print(f"  → {len(df)} products loaded{' (' + label + ')' if label else ''}\n")
    return df


# ---------------------------------------------------------------------------
# Diff report
# ---------------------------------------------------------------------------

def _values_equal(s1, s2):
    """Element-wise equality treating NaN == NaN as equal."""
    both_null = s1.isna() & s2.isna()
    return both_null | (s1 == s2)


def generate_report(df_submitted, df_sftp, submitted_path, diff_limit):
    """Return (report_sections, df_diffs, df_missing) where report_sections is a list of
    dicts, each with keys: 'level', 'summary' (dict), 'col_stats' (DataFrame or None).
    """
    sftp_ids = set(df_sftp["product-id"])

    shared_cols = [
        c for c in df_submitted.columns
        if c in df_sftp.columns and c not in ("product-id", "product-level")
    ]

    sections = []
    diff_rows = []
    missing_ids = []

    for level in PRODUCT_LEVELS:
        sub = df_submitted[df_submitted["product-level"] == level].copy()
        total = len(sub)
        if total == 0:
            continue

        found_mask = sub["product-id"].isin(sftp_ids)
        found = int(found_mask.sum())
        missing = total - found

        missing_ids.extend(sub[~found_mask]["product-id"].tolist())

        summary = {
            "level": level,
            "in_submitted": total,
            "found_in_sftp": found,
            "missing_from_sftp": missing,
        }

        common = sub[found_mask][["product-id"] + shared_cols]
        if common.empty:
            sections.append({"summary": summary, "col_stats": None})
            continue

        sftp_common = df_sftp[df_sftp["product-id"].isin(common["product-id"])][["product-id"] + shared_cols]
        merged = common.merge(sftp_common, on="product-id", suffixes=("_sub", "_sftp"))

        col_stat_rows = []
        for col in shared_cols:
            sub_col = f"{col}_sub"
            sftp_col = f"{col}_sftp"
            if sub_col not in merged.columns or sftp_col not in merged.columns:
                continue
            eq = _values_equal(merged[sub_col], merged[sftp_col])
            same = int(eq.sum())
            diff = int((~eq).sum())
            if diff == 0:
                continue
            prefixed_col = f"{level}.{col}"
            col_stat_rows.append({
                "column": prefixed_col,
                "same": same,
                "diff": diff,
                "diff%": round(diff / found * 100, 1),
            })

            diff_rows_for_col = merged[~eq][["product-id", sub_col, sftp_col]]
            if diff_limit is not None:
                diff_rows_for_col = diff_rows_for_col.head(diff_limit)
            for _, r in diff_rows_for_col.iterrows():
                diff_rows.append({
                    "column": prefixed_col,
                    "product-id": r["product-id"],
                    "comestri-value": r[sub_col],
                    "sftp-value": r[sftp_col],
                })

        sections.append({
            "summary": summary,
            "col_stats": pd.DataFrame(col_stat_rows) if col_stat_rows else None,
        })

    df_missing = df_submitted[df_submitted["product-id"].isin(missing_ids)][["product-id", "product-level"]].copy()

    all_col_stats = pd.concat(
        [sec["col_stats"] for sec in sections if sec["col_stats"] is not None],
        ignore_index=True,
    ) if any(sec["col_stats"] is not None for sec in sections) else pd.DataFrame()

    return sections, pd.DataFrame(diff_rows), df_missing, all_col_stats


def write_excel(sections, df_diffs, df_missing, submitted_path, wb_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    SECTION_FILL = PatternFill("solid", fgColor="2E75B6")
    ALT_FILL = PatternFill("solid", fgColor="D6E4F0")
    WHITE_FONT = Font(color="FFFFFF", bold=True)
    BOLD = Font(bold=True)

    wb = Workbook()

    # ── Tab 1: Report ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Report"

    def write_row(ws, row_idx, values, font=None, fill=None, alignment=None):
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment

    row = 1
    ws.cell(row=row, column=1, value="DIFF REPORT").font = Font(bold=True, size=14)
    row += 1
    ws.cell(row=row, column=1, value="Generated")
    ws.cell(row=row, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    row += 1
    ws.cell(row=row, column=1, value="Comestri export")
    ws.cell(row=row, column=2, value=submitted_path)
    row += 1
    ws.cell(row=row, column=1, value="SFTP full files")
    ws.cell(row=row, column=2, value=", ".join(os.path.basename(p) for p in SFTP_FULL_PATHS))
    row += 2

    # Product set summary table
    ws.cell(row=row, column=1, value="PRODUCT SET SUMMARY").font = BOLD
    row += 1
    write_row(ws, row, ["Level", "In Comestri Export", "Found in SFTP", "Missing from SFTP", "Missing %"],
              font=WHITE_FONT, fill=HEADER_FILL)
    row += 1
    for i, sec in enumerate(sections):
        s = sec["summary"]
        fill = ALT_FILL if i % 2 == 0 else None
        missing_pct = s["missing_from_sftp"] / s["in_submitted"] if s["in_submitted"] else 0
        write_row(ws, row, [s["level"], s["in_submitted"], s["found_in_sftp"], s["missing_from_sftp"], missing_pct], fill=fill)
        ws.cell(row=row, column=5).number_format = "0.0%"
        row += 1
    row += 1

    # Column stats per level
    ws.cell(row=row, column=1, value="COLUMN STATS (products present in both Comestri export and SFTP)").font = BOLD
    row += 1

    for sec in sections:
        s = sec["summary"]
        # Section heading
        write_row(ws, row,
                  [f"{s['level'].upper()}  —  {s['found_in_sftp']} common products"],
                  font=WHITE_FONT, fill=SECTION_FILL)
        row += 1

        if sec["col_stats"] is None or sec["col_stats"].empty:
            ws.cell(row=row, column=1, value="No differences found or no common products.")
            row += 2
            continue

        write_row(ws, row, ["Column", "Same", "Diff", "Diff%"],
                  font=WHITE_FONT, fill=HEADER_FILL)
        row += 1
        for i, (_, r) in enumerate(sec["col_stats"].iterrows()):
            fill = ALT_FILL if i % 2 == 0 else None
            write_row(ws, row, [r["column"], r["same"], r["diff"], r["diff%"] / 100],
                      fill=fill)
            ws.cell(row=row, column=4).number_format = "0.0%"
            row += 1
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 55
    for col in ["B", "C", "D"]:
        ws.column_dimensions[col].width = 18

    # ── Tab 2: Diff Details ──────────────────────────────────────────
    ws2 = wb.create_sheet("Diff Details")
    if df_diffs.empty:
        ws2.cell(row=1, column=1, value="No differences found.")
    else:
        write_row(ws2, 1, list(df_diffs.columns), font=WHITE_FONT, fill=HEADER_FILL)
        for i, (_, r) in enumerate(df_diffs.iterrows(), 2):
            fill = ALT_FILL if i % 2 == 0 else None
            write_row(ws2, i, list(r), fill=fill)
        ws2.column_dimensions["A"].width = 55
        ws2.column_dimensions["B"].width = 20
        ws2.column_dimensions["C"].width = 40
        ws2.column_dimensions["D"].width = 40

    # ── Tab 3: Missing Products ──────────────────────────────────────
    ws3 = wb.create_sheet("Missing Products")
    if df_missing.empty:
        ws3.cell(row=1, column=1, value="No missing products.")
    else:
        write_row(ws3, 1, list(df_missing.columns), font=WHITE_FONT, fill=HEADER_FILL)
        for i, (_, r) in enumerate(df_missing.iterrows(), 2):
            fill = ALT_FILL if i % 2 == 0 else None
            write_row(ws3, i, list(r), fill=fill)
        for col_idx, _ in enumerate(df_missing.columns, 1):
            ws3.column_dimensions[get_column_letter(col_idx)].width = 20

    wb.save(wb_path)



# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

DOWNLOADS_DIR = os.path.expanduser("~/Downloads")
SFTP_REPORT_DIR = "/incoming/diff_reports"
SFTP_SHEET_DATA_DIR = "/incoming/diff_reports/google_sheet_data"



def main():
    zip_path = input("Enter Comestri export file path: ").strip()
    diff_limit = input("Max differences per column in diff details (press Enter for no limit): ").strip()
    diff_limit = int(diff_limit) if diff_limit.isdigit() else None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_filename = f"report_{timestamp}.xlsx"
    wb_path = os.path.join(DOWNLOADS_DIR, report_filename)

    # Load submitted file
    print("\nLoading Comestri export ...")
    xmls = get_xml_buffers(zip_path)
    if not xmls:
        print("No XML files found.")
        return
    df_submitted = load_table(xmls, label="Comestri export")

    # Load SFTP full files
    print("\nLoading SFTP full files ...")
    sftp_chunks = []
    for remote_path in SFTP_FULL_PATHS:
        try:
            xml_buf = download_from_sftp(remote_path)
            name = os.path.basename(remote_path)
            if name.endswith(".gz"):
                name = name[:-3]
            sftp_chunks.append(pd.DataFrame(stream_products(xml_buf)))
            print(f"    → loaded {sftp_chunks[-1].shape[0]} products")
        except Exception as e:
            print(f"  [error] {remote_path}: {e}")

    if not sftp_chunks:
        print("No SFTP data loaded — cannot generate report.")
        return
    df_sftp = pd.concat(sftp_chunks, ignore_index=True)
    print(f"\nSFTP total: {len(df_sftp)} products")

    # Generate and write workbook
    print("\nGenerating report ...")
    sections, df_diffs, df_missing, df_col_stats = generate_report(df_submitted, df_sftp, zip_path, diff_limit)
    write_excel(sections, df_diffs, df_missing, zip_path, wb_path)

    col_stats_filename = f"column_stats_{timestamp}.csv"
    diff_details_filename = f"diff_details_{timestamp}.csv"

    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tmp:
        df_col_stats.to_csv(tmp.name, index=False)
        col_stats_tmp = tmp.name

    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False) as tmp:
        df_diffs.to_csv(tmp.name, index=False)
        diff_details_tmp = tmp.name

    # Upload to SFTP
    uploads = [
        (wb_path, SFTP_REPORT_DIR, report_filename),
        (col_stats_tmp, SFTP_SHEET_DATA_DIR, col_stats_filename),
        (diff_details_tmp, SFTP_SHEET_DATA_DIR, diff_details_filename),
    ]
    print(f"\nUploading to SFTP ...")
    for local_path, remote_dir, filename in uploads:
        remote_path_full = f"{remote_dir}/{filename}"
        try:
            transport = paramiko.Transport((os.environ["SFTP_HOST"], int(os.environ.get("SFTP_PORT", 22))))
            transport.connect(username=os.environ["SFTP_USERNAME"], password=os.environ["SFTP_PASSWORD"])
            sftp = paramiko.SFTPClient.from_transport(transport)
            sftp.put(local_path, remote_path_full)
            sftp.close()
            transport.close()
            print(f"  {remote_path_full}")
        except Exception as e:
            print(f"  [error] {remote_path_full}: {e}")

    os.unlink(col_stats_tmp)
    os.unlink(diff_details_tmp)

    # Print summary to terminal
    for sec in sections:
        s = sec["summary"]
        print(f"\n  {s['level'].upper()}: {s['found_in_sftp']}/{s['in_submitted']} found, "
              f"{s['missing_from_sftp']} missing")
        if sec["col_stats"] is not None and not sec["col_stats"].empty:
            print(f"  {len(sec['col_stats'])} column(s) with differences")

    print(f"\nWorkbook       → {wb_path}")
    print(f"  Tab 1: Report")
    print(f"  Tab 2: Diff Details ({len(df_diffs)} rows)")
    print(f"  Tab 3: Missing Products ({len(df_missing)} products)")
    print(f"Column stats   → {SFTP_SHEET_DATA_DIR}/{col_stats_filename} ({len(df_col_stats)} rows)")
    print(f"Diff details   → {SFTP_SHEET_DATA_DIR}/{diff_details_filename} ({len(df_diffs)} rows)")


if __name__ == "__main__":
    main()
